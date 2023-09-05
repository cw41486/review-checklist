import sys
import time
from tkinter import Tk
from openpyxl import load_workbook
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import asksaveasfile
from openpyxl.cell import MergedCell
import Compliance
import Specification
import Templates

#   method used to determine if a cell is merged.
def parent_of_merged_cell(cell):
    """ Find the parent of the merged cell by iterating through the range of merged cells """
    sheet = cell.parent
    child_coord = cell.coordinate

    # Note: if there are many merged cells in a large spreadsheet, this may become inefficient
    for merged in sheet.merged_cells.ranges:
        if child_coord in merged:
            return merged.start_cell.coordinate
    return None

Compliance.openComplianceWindow()     # open options window

#   Variables
comply_failures = []           # list of all failures
ecu = Compliance.ecu_entry
variant = Compliance.variant_entry
supplier = Compliance.supplier_entry
protocol = Compliance.protocol_entry
fullName = Compliance.name
mdp_version = Compliance.mdp_version
cda_version = Compliance.cda_version
efdFiles = ""
spec = None
dids = None
mandatoryDids = None
templates = None
maxDidSize = None
app_failures = {}
boot_failures = {}

#   Get Specifications.

if(protocol == "CS.00101"):
    spec = Specification.UDS_2()
    templates = Templates.Uds2Templates()
    dids = spec.dids
    mandatoryDids = spec.mandatoryDids
elif(protocol == 'CS-11825'):
    spec = Specification.Uds()
    dids = spec.dids
elif(protocol.startswith('CS.00021')):
    spec = Specification.Fiat()
    dids = spec.dids
else:
    pass

#   Check if main window was closed
if(ecu == "" and variant == "" and supplier == "" and protocol == ""):
    pass

else:
    #   Load a file
    root = Tk()
    root.withdraw()
    root.update()
    filename = askopenfilename(title="Load the Checklist")
    root.destroy()

    #   Load workbook.
    wb = load_workbook(filename)
    sheetNames = wb.sheetnames

    #   Open prompt to select the sheetname.
    Compliance.openExcelSheetSelectWindow(sheetNames)

    #   get the sheet name
    sheet = Compliance.excelSheet
    ws = wb[sheet]

    #   Grab data from worksheet.
    sheet_dids = ws['A9':'A37']
    did_names = ws['B9':'B32']
    atlantis_dids = ws['C33':'C37']
    app_response = ws['K9':'K37']
    boot_response = ws['L9':'L37']
    checklist_protocols = ws['D8':'J8']
    efdFiles = ws['D2'].value




    #   Lists for did data
    did_list = []
    protocol_did_list = []
    name_list = []
    app_list = []
    boot_list = []
    printedAppList = []
    printedBootList = []
    protocol_list = []

    for cell in sheet_dids:
        for x in cell:
            did_list.append(x.value)

    # Starting indexes
    column_index = 7
    did_index = -3
    name_index = -1

#   Search the checklist for 'x' or 'o' and grab the data.

    for cell in checklist_protocols:
        for i in cell:      # i = Cell object
            if(protocol in i.value):       #   If the selected protocol is detected, search that column and row for all necessary data.
                row_index = 1
                for x in range(0, len(did_list)):
                    try:        #  search the column for 'x' or 'o' and grab the data.
                        if(i.offset(row_index).value.startswith('X') or i.offset(row_index).value.startswith('O')):
                            app_list.append(i.offset(row_index,column_index).value)
                            boot_list.append(i.offset(row_index,column_index + 1).value)
                            protocol_did_list.append(i.offset(row_index,did_index).value)
                            if(isinstance(i.offset(row_index,name_index),MergedCell)):
                                val = parent_of_merged_cell(i.offset(row_index,name_index))
                                parent = i.offset(row_index,name_index).parent[val]
                                name_list.append(parent.value)
                            else:
                                name_list.append(i.offset(row_index,name_index).value)
                        row_index += 1
                    except AttributeError as ae:
                        row_index += 1
            column_index -= 1
            did_index -= 1
            name_index -= 1


    #   Check that all dids were collected.

    keySearch = []

    for x in protocol_did_list:
        keyDid = x[3:8].upper()
        keySearch.append(keyDid)


    notLocatedDids = []

    for k in dids.keys():
        if k in keySearch:
            pass
        else:
            notLocatedDids.append(k)

    if len(notLocatedDids) > 0:
        Compliance.infoWindow('Missing DIDs',f'Could not locate the following DIDs\n{notLocatedDids}\nPlease check the spreadsheet.')

    #   Create a new list filtering the Excel app and boot lists.

    for x in app_list:
        try:
            x = x.upper()
            new_spec = x.split(' ')
            remove_index = new_spec.index('62')
            for i in range(remove_index):
                del new_spec[0]

            for r in range(len(new_spec)):
                try:
                    new_spec.remove('')
                except ValueError:
                    None
                try:
                    new_spec.remove(' ')
                except ValueError:
                    None
                try:
                    new_spec.remove('CC')
                except ValueError:
                    None
                try:
                    new_spec.remove('cc')
                except ValueError:
                    None
            newString = ""
            for y in new_spec:
                newString += y + " "
            printedAppList.append(newString)

        except ValueError:
            printedAppList.append(x)

        except AttributeError:
            printedAppList.append("")

    for x in boot_list:
        try:
            x = x.upper()
            new_spec = x.split(' ')
            remove_index = new_spec.index('62')
            for i in range(remove_index):
                del new_spec[0]

            for r in range(len(new_spec)):
                try:
                    new_spec.remove('')
                except ValueError:
                    None
                try:
                    new_spec.remove(' ')
                except ValueError:
                    None
                try:
                    new_spec.remove('CC')
                except ValueError:
                    None
                try:
                    new_spec.remove('cc')
                except ValueError:
                    None

            newString = ""
            for y in new_spec:
                newString += y + " "
            printedBootList.append(newString)
        except ValueError:
            printedBootList.append(x)
        except AttributeError:
            printedBootList.append("")

    print(printedAppList)
    print(printedBootList)

    #   Prompt user to save file.
    Compliance.textFileGeneration()

    #   Save file as and store in a location of choice.
    root1 = Tk()
    root1.withdraw()
    root1.update()
    tf = asksaveasfile(
        mode='w',
        title="Save generated text file",
        initialfile=f"{ecu}_var{variant}_prtcl_cmply_{time.strftime('%d-%b-%y_%H-%M-%S')}.txt",
        defaultextension=".txt"
    )
    root1.destroy()

    #   Write data to the open file
    tf.write(f"""ECU:\t{ecu}
Variant:\t{variant}
Supplier:\t{supplier}
Date:\t{time.strftime('%d-%b-%Y')}\n
    """)

    #   Display values that are identical.

    if(len(printedAppList) != len(printedBootList)):
        Compliance.infoWindow('Error', 'Issue with data parsed from Excel file.  Please check the spreadsheet for validity.')
    else:
        for i in range(len(printedAppList)):
            if(printedAppList[i] == printedBootList[i]):
                pass
            else:
                comply_failures.append(printedAppList[i])
                comply_failures.append(printedBootList[i])

#   Check app and bootloader modes.
#
#   Application Mode
    for x in app_list:
        try:
            x = x.lower()
            startIndex = x.index('62')+3
            endIndex = startIndex + 5
            key_index = x[startIndex:endIndex]  # grab the did to search Dict.
            # key_index = x[3:8]
            spec_search = x.split(' ')  # split list into individual byte
            remove_index = spec_search.index('62') + 3
            for r in range(remove_index):
                try:
                    # spec_search.pop(0)  # Remove the start of the response.
                    del spec_search[0]
                except IndexError:
                    break
            for r in range(len(spec_search)):
                try:
                    spec_search.remove('')
                except ValueError:
                    None
                try:
                    spec_search.remove(' ')
                except ValueError:
                    None
                try:
                    spec_search.remove('CC')
                except ValueError:
                    None
                try:
                    spec_search.remove('cc')
                except ValueError:
                    None



            # # print(f'Spec - {spec_search}')
            if key_index.upper() in dids.keys():
                # # print(f"found key {key_index}")
                if isinstance(dids.get(key_index.upper()), list):
                    byte_num = 3
                    # # print(f'key_index {key_index}')
                    for did in spec_search:
                        if did in dids.get(key_index.upper()):
                            # # print(did)
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                app_failures.get(key_index.upper()).update({byte_num: did})
                                # app_failures.get(key_index.upper()).append(did) #   If key is already created, append to the list.
                            else:
                                # app_failures.update({key_index.upper():[did]})
                                app_failures.update({key_index.upper(): {byte_num: did}})  # add a new key to app_failures.
                        byte_num += 1

                else:
                    byte_num = 3
                    #   Check DIDs that could possibly have more than 1 LB of data
                    #   For F180, F181, F182.
                    if ((key_index.upper() == 'F1 80' or key_index.upper() == 'F1 81' or key_index.upper() == 'F1 82')  and protocol == 'CS.00101'):
                        numOfBlocks = int(spec_search[0],16)
                        #numOfBlocks = 1
                        byte_index = 3
                        total_bytes = 13 * numOfBlocks
                        for count in range(1):
                            for did in spec_search:
                                try:
                                    #print(str(byte_num) + " byte number")
                                    if did in dids.get(key_index.upper()).get(byte_num):
                                        #print(did + " byte number is within spec")
                                        pass
                                    else:
                                        if key_index.upper() in app_failures.keys():
                                            # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                            app_failures.get(key_index.upper()).update({byte_index: did})
                                        else:
                                            # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                            app_failures.update({key_index.upper(): {byte_index: did}})
                                    if(byte_num > 15):
                                        byte_num = 4
                                    else:
                                        byte_num += 1

                                    byte_index += 1
                                except TypeError:
                                    print("Type Error Detected")
                        if (byte_index - 4) != total_bytes:
                            actualSize = (byte_index - 4) / 13
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update({'NUMBER OF LBs': f'Shows data for {int(actualSize)} LBs but byte 3 shows {numOfBlocks} LBs'})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update({key_index.upper(): {'NUMBER OF LBs': f'Shows data for {int(actualSize)} LBs but byte 3 shows {numOfBlocks} LBs'}})


                    else:
                        for byte, resp in dids.get(key_index.upper()).items():
                            for did in spec_search:
                                if did in resp:
                                    byte_num += 1
                                else:
                                    if key_index.upper() in app_failures.keys():
                                        # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                        app_failures.get(key_index.upper()).update({byte_num: did})
                                    else:
                                        # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                        app_failures.update({key_index.upper(): {byte_num: did}})
                                    byte_num += 1
                                spec_search.pop(0)
                                break

            elif (key_index.upper() == 'F1 A0' and protocol == 'CS.00101'):     #  Check F1 A0 for UDS +2.
                byte_num = 3
                for did in spec_search:
                    print(byte_num)
                    if (byte_num >= 3 and byte_num <= 13):  # Check F187.
                        # print('F1 87')
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 14 and byte_num <= 23):  # Check F132
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 24 and byte_num <= 34):  # Check F191
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 35 and byte_num <= 45):  # Check F188
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 46 and byte_num <= 55):  # Check F18B
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})


                    elif (byte_num >= 56 and byte_num <= 65):  # Check F18A
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})
                    else:
                        if (did in spec.getFullHexRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    byte_num += 1

            elif (key_index.upper() == 'F1 A0' and protocol.startswith('CS.00021')):  # Check F1 A0 for UDS +2.
                byte_num = 3
                for did in spec_search:
                    print(byte_num)
                    if (byte_num >= 3 and byte_num <= 13):  # Check F187.
                        # print('F1 87')
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 14 and byte_num <= 30):  # Check F190
                        if (did in spec.getFullHexRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 31 and byte_num <= 41):  # Check F192
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 42 and byte_num <= 42):  # Check F193
                        if (did in spec.getFullHexRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 43 and byte_num <= 53):  # Check F194
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})


                    elif (byte_num >= 54 and byte_num <= 55):  # Check F195
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 56 and byte_num <= 61):  # Check F196
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 62 and byte_num <= 66):  # Check F1A5
                        if (did in spec.getFullHexRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    else:
                        if (did in spec.getFullHexRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update({key_index.upper(): {byte_num: did}})

                    byte_num += 1

        except AttributeError:
            pass
        except ValueError:
            print("Value error detected")

#   Bootloader
    for x in boot_list:
        try:
            x = x.lower()
            startIndex = x.index('62')+3
            endIndex = startIndex+5
            key_index = x[startIndex:endIndex]
            # key_index = x[3:8]  # grab the did to search Dict.
            spec_search = x.split(' ')  # split list into individual byte
            remove_index = spec_search.index('62') + 3
            for r in range(remove_index):
                try:
                    # spec_search.pop(0)  # Remove the start of the response.
                    del spec_search[0]
                except IndexError:
                    break

            # print(spec_search)

            for r in range(len(spec_search)):
                try:
                    spec_search.remove('')
                except ValueError:
                    None
                try:
                    spec_search.remove(' ')
                except ValueError:
                    None
                try:
                    spec_search.remove('CC')
                except ValueError:
                    None
                try:
                    spec_search.remove('cc')
                except ValueError:
                    None

            if key_index.upper() in dids.keys():
                # print(f"found key {key_index}")
                if isinstance(dids.get(key_index.upper()), list):
                    byte_num = 3
                    for did in spec_search:
                        if did in dids.get(key_index.upper()):
                            pass
                        else:
                            if key_index.upper() in boot_failures.keys():
                                boot_failures.get(key_index.upper()).update({byte_num: did})
                                # boot_failures.get(key_index.upper()).append(did) #   If key is already created, append to the list.
                            else:
                                # boot_failures.update({key_index.upper():[did]})
                                boot_failures.update(
                                    {key_index.upper(): {byte_num: did}})  # add a new key to boot_failures.
                        byte_num += 1

                else:
                    byte_num = 3
                    # for did in spec_search:

                    if (key_index.upper() == 'F1 80' and protocol == 'CS.00101'):
                        numOfBlocks = int(spec_search[0],16)
                        #numOfBlocks =
                        byte_index = 3
                        total_bytes = 13 * numOfBlocks
                        for count in range(1):
                            for did in spec_search:
                                try:
                                    #print(str(byte_num) + " byte number")
                                    if did in dids.get(key_index.upper()).get(byte_num):
                                        #print(did + " byte number is within spec")
                                        pass
                                    else:
                                        if key_index.upper() in boot_failures.keys():
                                            # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                            boot_failures.get(key_index.upper()).update({byte_index: did})
                                        else:
                                            # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                            boot_failures.update({key_index.upper(): {byte_index: did}})
                                    if(byte_num > 15):
                                        byte_num = 4
                                    else:
                                        byte_num += 1
                                    byte_index += 1
                                except TypeError:
                                    print("Type Error Detected")

                        if (byte_index - 4) != total_bytes:
                            actualSize = (byte_index - 4) / 13
                            print(str(int(actualSize))+"  Actual data shown")
                            if key_index.upper() in boot_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                boot_failures.get(key_index.upper()).update({'NUMBER OF LBs': f'Shows data for {int(actualSize)} LBs but byte 3 shows {numOfBlocks} LBs'})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                boot_failures.update({key_index.upper(): {'NUMBER OF LBs': f'Shows data for {int(actualSize)} LBs but byte 3 shows {numOfBlocks} LBs'}})

                    else:
                        for byte, resp in dids.get(key_index.upper()).items():
                            for did in spec_search:
                                if did in resp:
                                    byte_num += 1
                                else:
                                    if key_index.upper() in boot_failures.keys():
                                        # boot_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                        boot_failures.get(key_index.upper()).update({byte_num: did})
                                    else:
                                        # boot_failures.update({key_index.upper(): [did]})  # add a new key to boot_failures.
                                        boot_failures.update({key_index.upper(): {byte_num: did}})
                                    byte_num += 1
                                spec_search.pop(0)
                                break

            elif (key_index.upper() == 'F1 A0' and protocol == 'CS.00101'):
                byte_num = 3
                for did in spec_search:
                    if (byte_num >= 3 and byte_num <= 13):  # Check F187.
                        # print('F1 87')
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 14 and byte_num <= 23):  # Check F132
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 24 and byte_num <= 34):  # Check F191
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 35 and byte_num <= 45):  # Check F188
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 46 and byte_num <= 55):  # Check F18B
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})


                    elif (byte_num >= 56 and byte_num <= 65):  # Check F18A
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})
                    else:
                        if (did in spec.getFullHexRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})
                    byte_num += 1

            elif (key_index.upper() == 'F1 A0' and protocol.startswith('CS.00021')):  # Check F1 A0 for UDS +2.
                byte_num = 3
                for did in spec_search:
                    if (byte_num >= 3 and byte_num <= 13):  # Check F187.
                        # print('F1 87')
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 14 and byte_num <= 30):  # Check F190
                        if (did in spec.getFullHexRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 31 and byte_num <= 41):  # Check F192
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 42 and byte_num <= 42):  # Check F193
                        if (did in spec.getFullHexRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 43 and byte_num <= 53):  # Check F194
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})


                    elif (byte_num >= 54 and byte_num <= 55):  # Check F195
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 56 and byte_num <= 61):  # Check F196
                        if (did in spec.getValidAsciiRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    elif (byte_num >= 62 and byte_num <= 66):  # Check F1A5
                        if (did in spec.getFullHexRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update(
                                    {key_index.upper(): {byte_num: did}})

                    else:
                        if (did in spec.getFullHexRange()):
                            pass
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update(
                                    {byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update({key_index.upper(): {byte_num: did}})

                    byte_num += 1

        except AttributeError:
            pass
        except ValueError:
            print("Value error detected")

#    if(len(comply_failures) > 0):
#        tf.write(f"""
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
#THE FOLLOWING WERE DETECTED NOT TO BE IDENTICAL!  PLEASE CHECK.
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
#""")
#        print(comply_failures)
#        for i in comply_failures:
#            tf.write(f"""
#{i}
#""")
#   Write application responses not within spec.
    if len(app_failures.keys()) > 0:
        tf.write(f"""
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
APPLICATION MODE RESPONSES THAT ARE NOT WITHIN SPECIFICATIONS
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
""")
        index = 1
        print(app_failures)
        for k, v in app_failures.items():
            # print(k, v)
            tf.write(f"""
{index}. {k}:         
""")
            for byte,data in v.items():
                tf.write(f"\tbyte {byte}: {data}\n")
            index += 1

#   Write bootloader responses not within spec
    if len(boot_failures.keys()) > 0:
        tf.write(f"""
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
BOOTLOADER MODE RESPONSES THAT ARE NOT WITHIN SPECIFICATIONS
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
""")
        index = 1
        print(boot_failures)
        for k, v in boot_failures.items():
            tf.write(f"""
{index}. {k}:
""")
            for byte, data in v.items():
                tf.write(f"\tbyte {byte}: {data}\n")
            index += 1

#   Write the DIDs from the lists.

    tf.write(f"""
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
{protocol} - DIDs
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    
    """)

    ecuPartNumber = ""
    swPartNumber = ""
    hwPartNumber = ""
    failNumber = 1          #  Increments with each failure detected.

    # print(printedAppList)
    # print(printedBootList)

    for x in range(0, len(protocol_did_list)):  # Loop through all Dids an only collect Dids with responses.
        try:
            alph = ord('a')
            #   Locate Part Numbers.
            if '62 F1 32' in printedAppList[x]:
                #print("found ECU PN")
                ecuPartNumber = printedAppList[x]
            elif '62 F1 22' in printedAppList[x]:
                #print("Found SW PN")
                swPartNumber = printedAppList[x]
            elif '62 F1 12' in printedAppList[x]:
                hwPartNumber = printedAppList[x]
                #print("Found HW PN")
            else:
                pass
            # print(protocol_did_list[x])
            #    Begin searching the did list with app/boot failure list.
            startIndex = protocol_did_list[x].index('22')+3
            endIndex = startIndex + 5
            didSearch = protocol_did_list[x][startIndex:endIndex]
            failed = []
            # print(didSearch)
      #      try:        #  Go through results an print templates (if necessary).
      #          if didSearch in app_failures.keys():
      #              tf.write(f"\n{failNumber}) ${protocol_did_list[x]}: {name_list[x]}\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\nApp Mode:\t{printedAppList[x]}\nBoot Mode:\t{printedBootList[x]}\n\n")
      #              tf.write(f'{failNumber} - {templates.appInvalidRanges(app_failures.get(didSearch))}\n\n')
      #              failNumber += 1
      #          elif didSearch in boot_failures.keys():
      #              tf.write(f"\n{failNumber}) ${protocol_did_list[x]}: {name_list[x]}\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\nApp Mode:\t{printedAppList[x]}\nBoot Mode:\t{printedBootList[x]}\n\n")
      #              failNumber += 1
#
      #          elif printedAppList[x] != printedBootList[x] and didSearch in mandatoryDids:
      #              tf.write(f"\n{failNumber}) ${protocol_did_list[x]}: {name_list[x]}\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\nApp Mode:\t{printedAppList[x]}\nBoot Mode:\t{printedBootList[x]}\n\n")
      #              tf.write(f'{failNumber} - {templates.appAndBootMode}\n\n')
      #              failNumber += 1
#
      #          else:
      #              tf.write(f"\n${protocol_did_list[x]}: {name_list[x]}\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\nApp Mode:\t{printedAppList[x]}\nBoot Mode:\t{printedBootList[x]}\n\n")

            try:
                if (didSearch in app_failures.keys() and didSearch in boot_failures.keys()):        #  Check if the did is in both App and boot mode with invalid data.
                    failed.append(f'{failNumber}{chr(alph)}.  {templates.appAndBootInvalidRanges(app_failures.get(didSearch),boot_failures.get(didSearch))}\n')
                    alph += 1
                elif (didSearch in app_failures.keys() and didSearch not in boot_failures.keys()):    #  If only in app mode with invalid data, print those results.
                    failed.append(f'{failNumber}{chr(alph)}.  {templates.appInvalidRanges(app_failures.get(didSearch),didSearch)}\n')
                    alph += 1
                elif (didSearch not in app_failures.keys() and didSearch in boot_failures.keys()):    #  If only in boot mode with invalid data, print those results.
                    failed.append(f'{failNumber}{chr(alph)}.  {templates.bootInvalidRanges(boot_failures.get(didSearch),didSearch)}\n')
                    alph += 1
                if (printedAppList[x] != printedBootList[x] and didSearch in mandatoryDids):        #  Check if the results are identical between app and boot.
                    if didSearch == 'F1 32' or didSearch == 'F1 87':
                        failed.append(f'{failNumber}{chr(alph)}. {templates.invalidPartNumber}\n')
                        alph += 1
                        failed.append(f'{failNumber}{chr(alph)}. {templates.onlySupportsOnePartNumber}\n')
                        alph += 1
                    if didSearch == 'F1 22':
                        failed.append(f'{failNumber}{chr(alph)}.  {templates.expectedResponse}')
                        alph += 1
                    failed.append(f'{failNumber}{chr(alph)}. {templates.appAndBootMode}\n')
                    alph += 1

            except:
                print(sys.exc_info()[1], sys.exc_info()[2])

            try:
                if len(failed) > 0:
                    tf.write(f"\n{failNumber} ${protocol_did_list[x]}: {name_list[x]}\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\nApp Mode:\t{printedAppList[x]}\nBoot Mode:\t{printedBootList[x]}\n\n")
                    for x in failed:
                        tf.write(x)
                    failNumber += 1
                else:
                    tf.write(f"\n${protocol_did_list[x]}: {name_list[x]}\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\nApp Mode:\t{printedAppList[x]}\nBoot Mode:\t{printedBootList[x]}\n\n")

            except:
                print(sys.exc_info()[1], sys.exc_info()[2])


        except Exception as e:
            print(sys.exc_info()[1], sys.exc_info()[2])
            comply_failures.append(e)

    if (protocol.startswith('CS.00021')):
        tf.write(f"""
Tested by:  {fullName}
Date: {time.strftime('%d-%b-%Y')}
PN Update: 
Test Environment: Bench Setup
PC CDA Build: {cda_version}
MDP - {mdp_version}
Flash Results: All flashes performed via CDA 
Vector HW: End-End Successful. Abort-Recovery Successful. 
MDP: End-End Successful. Abort-Recovery Successful. 8v & 16v Flash Successful


Associated checksum: N/A
EFD File Used: {efdFiles}
$F1 32 PN: {ecuPartNumber}
$F1 22 PN: {swPartNumber}
$F1 12 PN: {hwPartNumber}

Associated checksum: N/A
EFD File Used: 
$F1 32 PN: 
$F1 22 PN: 
$F1 12 PN: 

Part Number Update Testing: 

$F1 95- Beginning read = 
Interrupted at 99% = 
Flashed to 100% = 

Abort-Recovery Testing:

Interrupted at  % = Successful
Interrupted at % = Successful
Interrupted at % = Successful
Interrupted at % = Successful
Interrupted at % = Successful
Flashed to 100% = Flash Successful

	""")

    elif (protocol.startswith('CS.00101') or protocol == 'CS-11825' or protocol.startswith('CS.00053')):
        tf.write(f"""
Tested by:  {fullName}
Date: {time.strftime('%d-%b-%Y')}
PN Update: 
Test Environment: Bench Setup
PC CDA Build: {cda_version}
MDP - {mdp_version}
Flash Results: All flashes performed via CDA 
Vector HW: End-End Successful. Abort-Recovery Successful. 
MDP: End-End Successful. Abort-Recovery Successful. 8v & 16v Flash Successful

Associated checksum: N/A
EFD File Used: {efdFiles}
$F1 32 PN:  {ecuPartNumber}
$F1 22 PN:  {swPartNumber}
$F1 12 PN:  {hwPartNumber}

Associated checksum: N/A
EFD File Used: 
$F1 32 PN: 
$F1 22 PN: 
$F1 12 PN: 

Part Number Update Testing:  

$F1 32- Beginning read = 
Interrupted at 99% = 
Flashed to 100% = 

Abort-Recovery Testing:

Interrupted at  % = Successful
Interrupted at % = Successful
Interrupted at % = Successful
Interrupted at % = Successful
Interrupted at % = Successful
Flashed to 100% = Flash Successful
	""")

    tf.close()

        #   Process complete
Compliance.processCompleted()
