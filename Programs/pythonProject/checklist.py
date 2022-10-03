import time
from tkinter import Tk
from openpyxl import load_workbook
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import asksaveasfile
from openpyxl.cell import MergedCell
import Compliance
import Specification


#   method used to determine if a cell is merged.
def parent_of_merged_cell(cell):
    """ Find the parent of the merged cell by iterating through the range of merged cells """
    sheet = cell.parent
    child_coord = cell.coordinate

    # Note: if there are many merged cells in a large spreadsheet, this may become inefficient
    for merged in sheet.merged_cells.ranges:
        if child_coord in merged:
            return merged.start_cell.coordinate
    return None

Compliance.openComplianceWindow()     # open options window

#   Variables
comply_failures = []           # list of all failures
ecu = Compliance.ecu_entry
variant = Compliance.variant_entry
supplier = Compliance.supplier_entry
protocol = Compliance.protocol_entry
fullName = Compliance.name
mdp_version = Compliance.mdp_version
cda_version = Compliance.cda_version
spec = None
dids = None
app_failures = {}
boot_failures = {}

#   Get Specifications.

if(protocol == "CS.00101"):
    spec = Specification.UDS_2()
    dids = spec.dids
elif(protocol == 'CS-11825'):
    spec = Specification.Uds()
    dids = spec.dids
elif(protocol.startswith('CS.00021')):
    spec = Specification.Fiat()
    dids = spec.dids
else:
    pass

#   Check if main window was closed
if(ecu == "" and variant == "" and supplier == "" and protocol == ""):
    pass

else:
    #   Load a file
    root = Tk()
    root.withdraw()
    root.update()
    filename = askopenfilename(title="Load the Checklist")
    root.destroy()

    #   Load workbook.
    wb = load_workbook(filename)
    sheetNames = wb.sheetnames

    #   Open prompt to select the sheetname.
    Compliance.openExcelSheetSelectWindow(sheetNames)

    #   get the sheet name
    sheet = Compliance.excelSheet
    ws = wb[sheet]

    #   Grab data from worksheet.
    sheet_dids = ws['A9':'A37']
    did_names = ws['B9':'B32']
    atlantis_dids = ws['C33':'C37']
    app_response = ws['K9':'K37']
    boot_response = ws['L9':'L37']
    checklist_protocols = ws['D8':'J8']

    #   Lists for did data
    did_list = []
    protocol_did_list = []
    name_list = []
    app_list = []
    boot_list = []
    protocol_list = []

    for cell in sheet_dids:
        for x in cell:
            did_list.append(x.value)

    # Starting indexes
    column_index = 7
    did_index = -3
    name_index = -1

#   Search the checklist for 'x' or 'o' and grab the data.

    for cell in checklist_protocols:
        for i in cell:      # i = Cell object
            if(protocol in i.value):       #   If the selected protocol is detected, search that column and row for all necessary data.
                row_index = 1
                for x in range(0,len(did_list)):
                    try:        #  search the column for 'x' or 'o' and grab the data.
                        if(i.offset(row_index).value.startswith('X') or i.offset(row_index).value.startswith('O')):
                            app_list.append(i.offset(row_index,column_index).value)
                            boot_list.append(i.offset(row_index,column_index + 1).value)
                            protocol_did_list.append(i.offset(row_index,did_index).value)
                            if(isinstance(i.offset(row_index,name_index),MergedCell)):
                                val = parent_of_merged_cell(i.offset(row_index,name_index))
                                parent = i.offset(row_index,name_index).parent[val]
                                name_list.append(parent.value)
                            else:
                                name_list.append(i.offset(row_index,name_index).value)
#                            if(i.offset(row_index,name_index).value == 'None'):
#                                name_index -= 1
#                            name_list.append(i.offset(row_index,name_index).value)
                        row_index += 1
                    except AttributeError as ae:
                        row_index += 1
            column_index -= 1
            did_index -= 1
            name_index -= 1

    #   Prompt user to save file.
    Compliance.textFileGeneration()

    #   Save file as and store in a location of choice.
    root1 = Tk()
    root1.withdraw()
    root1.update()
    tf = asksaveasfile(
        mode='w',
        title="Save generated text file",
        initialfile=f"{ecu}_var{variant}_prtcl_cmply_{time.strftime('%d-%b-%y_%H-%M-%S')}.txt",
        defaultextension=".txt"
    )
    root1.destroy()

    #   Write data to the open file
    tf.write(f"""ECU:\t{ecu}
Variant:\t{variant}
Supplier:\t{supplier}
Date:\t{time.strftime('%d-%b-%Y')}\n
    """)

    #   Display values that are identical.

    if(len(app_list) != len(boot_list)):
        Compliance.infoWindow('Error', 'Issue with data parsed from Excel file.  Please check the spreadsheet for validity.')
    else:
        for i in range(len(app_list)):
            if(app_list[i] == boot_list[i]):
                pass
            else:
                comply_failures.append(app_list[i])
                comply_failures.append(boot_list[i])

#   Check app and bootloader modes.

    for x in app_list:
        x = x.lower()
        key_index = x[3:8]  # grab the did to search Dict.
        spec_search = x.split(' ')  # split list into individual byte
        for r in range(3):
            try:
                spec_search.pop(0)  # Remove the start of the response.
            except IndexError:
                break
        for r in range(len(spec_search)):
            try:
                spec_search.remove('')
            except ValueError:
                None
        # # print(f'Spec - {spec_search}')
        if key_index.upper() in dids.keys():
            # # print(f"found key {key_index}")
            if isinstance(dids.get(key_index.upper()), list):
                byte_num = 3
                # # print(f'key_index {key_index}')
                for did in spec_search:
                    if did in dids.get(key_index.upper()):
                        # # print(did)
                        pass
                    else:
                        if key_index.upper() in app_failures.keys():
                            app_failures.get(key_index.upper()).update({byte_num: did})
                            # app_failures.get(key_index.upper()).append(did) #   If key is already created, append to the list.
                        else:
                            # app_failures.update({key_index.upper():[did]})
                            app_failures.update({key_index.upper(): {byte_num: did}})  # add a new key to app_failures.
                    byte_num += 1

            else:
                byte_num = 3
                # # print(f'key index {key_index}')
                # for did in spec_search:
                for byte, resp in dids.get(key_index.upper()).items():
                    for did in spec_search:
                        if did in resp:
                            byte_num += 1
                        else:
                            if key_index.upper() in app_failures.keys():
                                # app_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                app_failures.get(key_index.upper()).update({byte_num: did})
                            else:
                                # app_failures.update({key_index.upper(): [did]})  # add a new key to app_failures.
                                app_failures.update({key_index.upper(): {byte_num: did}})
                            byte_num += 1
                        spec_search.pop(0)
                        break



#   Bootloader
    for x in boot_list:
        x = x.lower()
        key_index = x[3:8]  # grab the did to search Dict.
        spec_search = x.split(' ')  # split list into individual byte
        for r in range(3):
            try:
                spec_search.pop(0)  # Remove the start of the response.
            except IndexError:
                break
        for r in range(len(spec_search)):
            try:
                spec_search.remove('')
            except ValueError:
                None

        if key_index.upper() in dids.keys():
            # print(f"found key {key_index}")
            if isinstance(dids.get(key_index.upper()), list):
                byte_num = 3
                for did in spec_search:
                    if did in dids.get(key_index.upper()):
                        pass
                    else:
                        if key_index.upper() in boot_failures.keys():
                            boot_failures.get(key_index.upper()).update({byte_num: did})
                            # boot_failures.get(key_index.upper()).append(did) #   If key is already created, append to the list.
                        else:
                            # boot_failures.update({key_index.upper():[did]})
                            boot_failures.update(
                                {key_index.upper(): {byte_num: did}})  # add a new key to boot_failures.
                    byte_num += 1

            else:
                byte_num = 3
                # for did in spec_search:
                for byte, resp in dids.get(key_index.upper()).items():
                    for did in spec_search:
                        if did in resp:
                            byte_num += 1
                        else:
                            if key_index.upper() in boot_failures.keys():
                                # boot_failures.get(key_index.upper()).append(did)  # If key is already created, append to the list.
                                boot_failures.get(key_index.upper()).update({byte_num: did})
                            else:
                                # boot_failures.update({key_index.upper(): [did]})  # add a new key to boot_failures.
                                boot_failures.update({key_index.upper(): {byte_num: did}})
                            byte_num += 1
                        spec_search.pop(0)
                        break

    if(len(comply_failures) > 0):
        tf.write(f"""
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
THE FOLLOWING WERE DETECTED NOT TO BE IDENTICAL!  PLEASE CHECK.
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
""")
        for i in comply_failures:
            tf.write(f"""
{i}
""")
#   Write application responses not within spec.
    if len(app_failures.keys()) > 0:
        tf.write(f"""
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
APPLICATION MODE RESPONSES THAT ARE NOT WITHIN SPECIFICATIONS
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
""")
        index = 1
        for k, v in app_failures.items():
            # print(k, v)
            tf.write(f"""
{index}. {k} - {v}
         
""")
            index += 1

#   Write bootloader responses not within spec
    if len(boot_failures.keys()) > 0:
        tf.write(f"""
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
BOOTLOADER MODE RESPONSES THAT ARE NOT WITHIN SPECIFICATIONS
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
""")
        index = 1
        for k, v in boot_failures.items():
            tf.write(f"""
{index}. {k} - {v}
""")
            index += 1

#   Write the DIDs from the lists.

    tf.write(f"""
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
{protocol} - DIDs
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    """)

    for x in range(0, len(protocol_did_list)):  # Loop through all Dids an only collect Dids with responses.
        try:
            tf.write(f"""
${protocol_did_list[x]}: {name_list[x]}
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
App Mode:\t{app_list[x]}
Boot Mode:\t{boot_list[x]}

""")
        except Exception as e:
            comply_failures.append(e)

    if (protocol.startswith('CS.00021')):
        tf.write(f"""
Tested by:  {fullName}
Date: {time.strftime('%d-%b-%Y')}
PN Update: 
Test Environment: Bench Setup
PC CDA Build: {cda_version}
MDP - {mdp_version}
Flash Results: All flashes performed via CDA 
Vector HW: End-End Successful. Abort-Recovery Successful. 
MDP: End-End Successful. Abort-Recovery Successful. 8v & 16v Flash Successful


Associated checksum: N/A
EFD File Used:
$F1 32 PN: 
$F1 22 PN: 
$F1 12 PN: 

Associated checksum: N/A
EFD File Used: 
$F1 32 PN: 
$F1 22 PN: 
$F1 12 PN: 

Part Number Update Testing: 

$F1 95- Beginning read = 
Interrupted at 99% = 
Flashed to 100% = 

Abort-Recovery Testing:

Interrupted at  % = Successful
Interrupted at % = Successful
Interrupted at % = Successful
Interrupted at % = Successful
Interrupted at % = Successful
Flashed to 100% = Flash Successful

	""")

    elif (protocol.startswith('CS.00101') or protocol == 'CS-11825' or protocol.startswith('CS.00053')):
        tf.write(f"""
Tested by:  {fullName}
Date: {time.strftime('%d-%b-%Y')}
PN Update: 
Test Environment: Bench Setup
PC CDA Build: {cda_version}
MDP - {mdp_version}
Flash Results: All flashes performed via CDA 
Vector HW: End-End Successful. Abort-Recovery Successful. 
MDP: End-End Successful. Abort-Recovery Successful. 8v & 16v Flash Successful

Associated checksum: N/A
EFD File Used:
$F1 32 PN: 
$F1 22 PN: 
$F1 12 PN: 

Associated checksum: N/A
EFD File Used: 
$F1 32 PN: 
$F1 22 PN: 
$F1 12 PN: 

Part Number Update Testing:  

$F1 32- Beginning read = 
Interrupted at 99% = 
Flashed to 100% = 

Abort-Recovery Testing:

Interrupted at  % = Successful
Interrupted at % = Successful
Interrupted at % = Successful
Interrupted at % = Successful
Interrupted at % = Successful
Flashed to 100% = Flash Successful
	""")

    tf.close()

        #   Process complete
Compliance.processCompleted()
